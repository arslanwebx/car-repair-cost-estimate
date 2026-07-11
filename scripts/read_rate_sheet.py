import sys
from openpyxl import load_workbook

path, sheet_name = sys.argv[1], sys.argv[2]
values = load_workbook(path, data_only=True, read_only=True)[sheet_name]
formulas = load_workbook(path, data_only=False, read_only=True)[sheet_name]
for value_row, formula_row in zip(values.iter_rows(), formulas.iter_rows()):
    cells = []
    for value_cell, formula_cell in zip(value_row, formula_row):
        value = value_cell.value
        formula = formula_cell.value
        if isinstance(formula, str) and formula.startswith("="):
            cells.append(f"{value!r} [{formula}]")
        else:
            cells.append(repr(value))
    if any(cell.value is not None for cell in value_row):
        print(f"{value_row[0].row}\t" + "\t".join(cells))
