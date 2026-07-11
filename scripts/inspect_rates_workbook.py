import json
import sys
from pathlib import Path
from openpyxl import load_workbook

path = Path(sys.argv[1])
book = load_workbook(path, data_only=False, read_only=False)
result = {"file": str(path), "sheets": []}
for sheet in book.worksheets:
    rows = []
    for row in sheet.iter_rows():
        values = [cell.value for cell in row]
        if any(value is not None for value in values):
            rows.append({"row": row[0].row, "values": values})
    result["sheets"].append({
        "title": sheet.title,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "merged_cells": [str(item) for item in sheet.merged_cells.ranges],
        "rows": rows,
    })
print(json.dumps(result, indent=2, default=str))
