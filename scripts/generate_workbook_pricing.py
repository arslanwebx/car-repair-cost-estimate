import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from openpyxl import load_workbook

source = Path(sys.argv[1])
target = Path(sys.argv[2])
book = load_workbook(source, data_only=True, read_only=True)

def clean(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value

def rows(sheet, start, end, columns):
    result = []
    for row in range(start, end + 1):
        result.append({name: clean(sheet.cell(row, column).value) for name, column in columns.items()})
    return result

model_sheet = book["Model_Inputs"]
model_inputs = {}
for row in range(5, 36):
    name = model_sheet.cell(row, 1).value
    if name:
        model_inputs[name] = clean(model_sheet.cell(row, 2).value)

usa = rows(book["USA_Rates"], 5, 55, {
    "state": 1, "abbreviation": 2, "region": 3, "laborMultiplier": 8,
    "partsMultiplier": 9, "zone": 10, "bodyRate": 11, "paintRate": 12,
    "frameRate": 13, "mechanicalRate": 14, "paintMaterialRate": 15,
    "scanFee": 16, "staticCalibration": 17, "dynamicCalibration": 18,
    "shopSuppliesRate": 19, "hazmatFee": 20, "hiddenReserve": 21,
})

operations = rows(book["Damage_Catalog"], 5, 77, {
    "code": 1, "category": 2, "area": 3, "visibleDamage": 4, "severity": 5,
    "strategy": 6, "bodyHours": 7, "refinishHours": 8, "frameHours": 9,
    "mechanicalHours": 10, "basePart": 11, "scanNeed": 12,
    "calibrationNeed": 13, "blendPanels": 14, "consumables": 15,
    "lowFactor": 16, "highFactor": 17, "notes": 18,
    "nationalMid": 20, "nationalLow": 21, "nationalHigh": 22,
})

vehicle_sheet = book["Vehicle_Adjustments"]
vehicle_classes = rows(vehicle_sheet, 5, 13, {"name": 1, "labor": 2, "parts": 3, "calibration": 4, "notes": 5})
powertrains = rows(vehicle_sheet, 5, 9, {"name": 8, "mechanicalLabor": 9, "parts": 10, "calibration": 11, "notes": 12})
materials = rows(vehicle_sheet, 5, 8, {"name": 14, "bodyLabor": 15, "parts": 16, "notes": 17})
paint_types = rows(vehicle_sheet, 18, 23, {"name": 1, "paintLabor": 2, "materials": 3, "notes": 4})
parts_sources = rows(vehicle_sheet, 18, 22, {"name": 6, "price": 7, "risk": 8, "notes": 9})
vehicle_ages = rows(vehicle_sheet, 18, 21, {"name": 11, "parts": 12, "labor": 13, "notes": 14})

global_sheet = book["Global_Markets"]
global_column_count = 19
headers = [global_sheet.cell(4, column).value for column in range(1, global_column_count + 1)]
global_markets = []
for row in range(5, 59):
    global_markets.append({str(header): clean(global_sheet.cell(row, column).value) for column, header in enumerate(headers, 1) if header})

config = {
    "version": "US-WB-2026.07-v1",
    "sourceWorkbook": source.name,
    "worksheetAudit": [{"name": sheet.title, "rows": sheet.max_row, "columns": sheet.max_column} for sheet in book.worksheets],
    "modelInputs": model_inputs,
    "usaRates": usa,
    "damageOperations": operations,
    "vehicleAdjustments": {
        "classes": vehicle_classes, "powertrains": powertrains, "materials": materials,
        "paintTypes": paint_types, "partsSources": parts_sources, "vehicleAges": vehicle_ages,
    },
    "globalMarkets": global_markets,
}
target.parent.mkdir(parents=True, exist_ok=True)
target.write_text(json.dumps(config, indent=2, ensure_ascii=False), encoding="utf-8")
print(f"Generated {target} with {len(usa)} state rows, {len(operations)} operations, and {len(global_markets)} global markets.")
